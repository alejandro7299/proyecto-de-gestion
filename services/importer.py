"""
Excel Importer - Modo streaming para bajo consumo de memoria
=============================================================
Usa openpyxl en modo read_only para procesar el Excel fila por fila
sin cargarlo todo en RAM. Soporta archivos de 10MB+ en servidores
con 512MB de RAM (Render free tier).
"""
 
import unicodedata
import re
import logging
from openpyxl import load_workbook
 
logger = logging.getLogger(__name__)
 
CANONICAL_COLUMNS = {
    "patente": "patente", "marca": "marca", "modelo": "modelo",
    "comprobante": "comprobante", "fecha": "fecha", "mes": "mes",
    "observacion": "observacion", "observación": "observacion",
    "detalle": "detalle", "accion": "accion", "acción": "accion",
    "subrubro": "subrubro", "insumo": "insumo",
    "nominsumo": "nominsumo", "nom_insumo": "nominsumo",
    "taller": "taller", "cantidad": "cantidad",
    "precio_unit": "precio_unit", "precio unit": "precio_unit",
    "preciounit": "precio_unit", "precio_unitario": "precio_unit",
    "costo": "costo", "centrocosto": "centrocosto",
    "centro_costo": "centrocosto", "operador": "operador",
    "n_ot": "n_ot", "not": "n_ot", "n_ot": "n_ot",
    "panol": "panol",
    "inicio_ot": "inicio_ot", "tecnico": "tecnico",
    "cumplida": "cumplida", "fin_ot": "fin_ot",
    "operacion": "operacion", "fletero": "fletero",
    "empresa": "empresa", "rubro": "rubro",
}
 
NUMERIC_COLUMNS = {"cantidad", "precio_unit", "costo"}
DEFAULT_FILL = "SIN SELECCIONAR"
BATCH_SIZE = 500
 
 
def normalize_col_name(name: str) -> str:
    if not isinstance(name, str):
        name = str(name)
    name = unicodedata.normalize("NFKD", name)
    name = "".join(c for c in name if not unicodedata.combining(c))
    name = name.lower().strip().replace("ñ", "n")
    name = re.sub(r"[\s/\\-]+", "_", name)
    name = re.sub(r"[^\w]", "", name)
    name = re.sub(r"_+", "_", name).strip("_")
    return name
 
 
def map_to_canonical(col: str) -> str:
    return CANONICAL_COLUMNS.get(normalize_col_name(col), normalize_col_name(col))
 
 
def clean_value(val, col_name: str):
    if val is None:
        return 0.0 if col_name in NUMERIC_COLUMNS else DEFAULT_FILL
    str_val = str(val).strip()
    if str_val.lower() in ("none", "nan", ""):
        return 0.0 if col_name in NUMERIC_COLUMNS else DEFAULT_FILL
    if col_name in NUMERIC_COLUMNS:
        try:
            return float(str_val.replace(",", "."))
        except (ValueError, AttributeError):
            return 0.0
    return str_val
 
 
def load_excel_robust(file_obj) -> tuple:
    """
    Carga el Excel en modo streaming (read_only=True).
    No carga todo en memoria - procesa fila por fila.
    """
    report = {
        "header_row": None,
        "rows_imported": 0,
        "rows_skipped": 0,
        "columns_mapped": {},
        "columns_unknown": [],
        "warnings": [],
    }
 
    wb = load_workbook(filename=file_obj, read_only=True, data_only=True)
    ws = wb.active
 
    header_idx = None
    canonical_cols = []
    rows_data = []
    row_num = 0
 
    for row in ws.iter_rows(values_only=True):
        row_num += 1
 
        if header_idx is None:
            for cell in row:
                if isinstance(cell, str) and "patente" in cell.lower().strip():
                    header_idx = row_num
                    report["header_row"] = row_num
                    canonical_cols = [
                        map_to_canonical(str(c)) if c is not None else None
                        for c in row
                    ]
                    report["columns_mapped"] = {
                        str(c): map_to_canonical(str(c))
                        for c in row if c is not None
                    }
                    break
            continue
 
        values = [v for v in row if v is not None and str(v).strip() not in ("", "None", "nan")]
        if len(values) < 3:
            report["rows_skipped"] += 1
            continue
 
        row_dict = {}
        for i, val in enumerate(row):
            if i >= len(canonical_cols):
                break
            col = canonical_cols[i]
            if col is None:
                continue
            row_dict[col] = clean_value(val, col)
 
        all_canonical = set(CANONICAL_COLUMNS.values())
        for col in all_canonical:
            if col not in row_dict:
                row_dict[col] = 0.0 if col in NUMERIC_COLUMNS else DEFAULT_FILL
 
        rows_data.append(row_dict)
 
    wb.close()
 
    if header_idx is None:
        raise ValueError(
            "No se encontró la columna 'Patente' en el archivo. "
            "Verificá que el Excel contenga la tabla de datos correcta."
        )
 
    report["rows_imported"] = len(rows_data)
    return rows_data, report
 
 
def insert_dataframe(conn, rows_data: list, report: dict) -> int:
    """Inserta en batches. Compatible con SQLite y PostgreSQL."""
    from db import USE_SQLITE
 
    if not rows_data:
        return 0
 
    cols = [c for c in rows_data[0].keys() if c not in ("id", "created_at")]
    ph = "?" if USE_SQLITE else "%s"
    placeholders = ", ".join([ph] * len(cols))
    sql = f"INSERT INTO flota ({', '.join(cols)}) VALUES ({placeholders})"
 
    total = 0
    cur = conn.cursor()
    for i in range(0, len(rows_data), BATCH_SIZE):
        batch = rows_data[i:i + BATCH_SIZE]
        batch_tuples = [tuple(row.get(c, DEFAULT_FILL) for c in cols) for row in batch]
        cur.executemany(sql, batch_tuples)
        conn.commit()
        total += len(batch)
    cur.close()
    return total
 